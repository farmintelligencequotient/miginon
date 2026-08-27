import csv
import re
from io import BytesIO

from django.http import HttpResponse


def _filename(report, ext):
    slug = re.sub(r'[^a-z0-9]+', '-', report['period_label'].lower()).strip('-')
    return f"{report['farm'].code}-{slug}-report.{ext}"


# --------------------------------------------------------------------- CSV

def export_csv(report):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{_filename(report, "csv")}"'
    writer = csv.writer(response)

    farm = report['farm']
    writer.writerow([f'{farm.name} ({farm.code}) - {report["period_label"]} Report'])
    writer.writerow([f'Period: {report["start_date"]} to {report["end_date"]}'])
    writer.writerow([f'Generated {report["generated_at"].strftime("%d %b %Y %H:%M")}'])
    writer.writerow([])

    writer.writerow(['MILK PRODUCTION'])
    writer.writerow(['Total liters', report['milk']['total_liters']])
    writer.writerow(['Avg liters per active cow', round(float(report['milk']['avg_per_cow']), 2)])
    writer.writerow([])
    writer.writerow(['By block'])
    writer.writerow(['Block', 'Liters'])
    for row in report['milk']['by_block']:
        writer.writerow([row['block__name'], row['liters']])
    writer.writerow([])
    writer.writerow(['By cow'])
    writer.writerow(['Tag', 'Name', 'Block', 'Liters'])
    for row in report['milk']['by_cow']:
        writer.writerow([row['cow__tag_id'], row['cow__name'] or '', row['cow__block__name'], row['liters']])
    writer.writerow([])
    writer.writerow(['Daily totals'])
    writer.writerow(['Date', 'Liters'])
    for row in report['milk']['daily']:
        writer.writerow([row['date'], row['liters']])
    writer.writerow([])

    writer.writerow(['FEEDING'])
    writer.writerow(['Total dairy meal (kg)', report['feeding']['total_dairy_meal_kg']])
    writer.writerow(['Total silage/hay (kg)', report['feeding']['total_silage_hay_kg']])
    writer.writerow([])
    writer.writerow(['By block'])
    writer.writerow(['Block', 'Dairy meal (kg)', 'Silage/Hay (kg)'])
    for row in report['feeding']['by_block']:
        writer.writerow([row['block__name'], row['meal'] or 0, row['silage'] or 0])
    writer.writerow([])
    writer.writerow(['Daily totals'])
    writer.writerow(['Date', 'Dairy meal (kg)', 'Silage/Hay (kg)'])
    for row in report['feeding']['daily']:
        writer.writerow([row['date'], row['meal'] or 0, row['silage'] or 0])
    writer.writerow([])

    writer.writerow(['HERD'])
    writer.writerow(['Active cows', report['herd']['active_cows']])
    writer.writerow([])
    writer.writerow(['New cows added'])
    writer.writerow(['Tag', 'Name', 'Block', 'Breed', 'Added on'])
    for cow in report['herd']['new_cows']:
        writer.writerow([cow.tag_id, cow.name, cow.block.name, cow.breed, cow.created_at.strftime('%Y-%m-%d')])
    writer.writerow([])
    writer.writerow(['Transfers'])
    writer.writerow(['Cow', 'From block', 'To block', 'Date', 'Note'])
    for t in report['herd']['transfers']:
        writer.writerow([
            t.cow.tag_id, t.from_block.name if t.from_block else '', t.to_block.name,
            t.transferred_at.strftime('%Y-%m-%d'), t.note,
        ])
    writer.writerow([])

    writer.writerow(['CROPS'])
    writer.writerow(['Total harvested (kg)', report['crops']['total_harvested_kg']])
    writer.writerow([])
    writer.writerow(['Date', 'Crop', 'Activity', 'Harvested (kg)', 'Notes'])
    for a in report['crops']['activities']:
        writer.writerow([
            a.date, a.crop.name, a.get_activity_type_display(),
            a.quantity_harvested_kg if a.quantity_harvested_kg is not None else '', a.notes,
        ])
    writer.writerow([])

    writer.writerow(['INVENTORY'])
    writer.writerow(['Movements'])
    writer.writerow(['Date', 'Item', 'Type', 'Quantity', 'Note'])
    for m in report['inventory']['movements']:
        writer.writerow([m.date, m.item.name, m.get_movement_type_display(), m.quantity, m.note])
    writer.writerow([])
    writer.writerow(['Current stock levels'])
    writer.writerow(['Item', 'Category', 'Stock', 'Unit', 'Reorder level', 'Low stock?'])
    for item in report['inventory']['levels']:
        writer.writerow([
            item.name, item.get_category_display(), item.current_stock, item.get_unit_display(),
            item.reorder_level, 'Yes' if item.is_low_stock else 'No',
        ])
    writer.writerow([])

    writer.writerow(['FINANCE'])
    writer.writerow(['Total income', report['finance']['total_income']])
    writer.writerow(['Total expense', report['finance']['total_expense']])
    writer.writerow(['Net', report['finance']['net']])
    writer.writerow([])
    writer.writerow(['By category'])
    writer.writerow(['Kind', 'Category', 'Amount'])
    for row in report['finance']['by_category']:
        writer.writerow([row['kind'], row['category'], row['amount']])
    writer.writerow([])
    writer.writerow(['Transactions'])
    writer.writerow(['Date', 'Kind', 'Category', 'Amount', 'Note'])
    for t in report['finance']['transactions']:
        writer.writerow([t.date, t.get_kind_display(), t.get_category_display(), t.amount, t.note])

    return response


# -------------------------------------------------------------------- XLSX

def export_xlsx(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='047857')
    title_font = Font(bold=True, size=14, color='065F46')

    def style_header(ws, row_idx, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = header_font
            cell.fill = header_fill

    def autosize(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 42)

    def section_table(ws, title, headers, rows):
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append(headers)
        style_header(ws, ws.max_row, len(headers))
        for row in rows:
            ws.append(row)
        ws.append([])

    farm = report['farm']

    ws = wb.active
    assert ws is not None
    ws.title = 'Summary'
    ws['A1'] = f'{farm.name} ({farm.code})'
    ws['A1'].font = title_font
    ws['A2'] = f'{report["period_label"]} report ({report["start_date"]} to {report["end_date"]})'
    ws['A3'] = f'Generated {report["generated_at"].strftime("%d %b %Y %H:%M")}'
    ws.append([])
    ws.append(['Metric', 'Value'])
    style_header(ws, ws.max_row, 2)
    ws.append(['Total milk (L)', float(report['milk']['total_liters'])])
    ws.append(['Avg milk per active cow (L)', round(float(report['milk']['avg_per_cow']), 2)])
    ws.append(['Total dairy meal (kg)', float(report['feeding']['total_dairy_meal_kg'])])
    ws.append(['Total silage/hay (kg)', float(report['feeding']['total_silage_hay_kg'])])
    ws.append(['Active cows', report['herd']['active_cows']])
    ws.append(['New cows added', len(report['herd']['new_cows'])])
    ws.append(['Cow transfers', len(report['herd']['transfers'])])
    ws.append(['Total harvested (kg)', float(report['crops']['total_harvested_kg'])])
    ws.append(['Total income', float(report['finance']['total_income'])])
    ws.append(['Total expense', float(report['finance']['total_expense'])])
    ws.append(['Net', float(report['finance']['net'])])
    autosize(ws)

    ws = wb.create_sheet('Milk')
    section_table(ws, 'By block', ['Block', 'Liters'],
                  [[r['block__name'], float(r['liters'])] for r in report['milk']['by_block']])
    section_table(ws, 'By cow', ['Tag', 'Name', 'Block', 'Liters'],
                  [[r['cow__tag_id'], r['cow__name'] or '', r['cow__block__name'], float(r['liters'])]
                   for r in report['milk']['by_cow']])
    section_table(ws, 'Daily totals', ['Date', 'Liters'],
                  [[r['date'].strftime('%Y-%m-%d'), float(r['liters'])] for r in report['milk']['daily']])
    autosize(ws)

    ws = wb.create_sheet('Feeding')
    section_table(ws, 'By block', ['Block', 'Dairy meal (kg)', 'Silage/Hay (kg)'],
                  [[r['block__name'], float(r['meal'] or 0), float(r['silage'] or 0)]
                   for r in report['feeding']['by_block']])
    section_table(ws, 'Daily totals', ['Date', 'Dairy meal (kg)', 'Silage/Hay (kg)'],
                  [[r['date'].strftime('%Y-%m-%d'), float(r['meal'] or 0), float(r['silage'] or 0)]
                   for r in report['feeding']['daily']])
    autosize(ws)

    ws = wb.create_sheet('Herd')
    section_table(ws, 'New cows added', ['Tag', 'Name', 'Block', 'Breed', 'Added on'],
                  [[c.tag_id, c.name or '', c.block.name, c.breed or '', c.created_at.strftime('%Y-%m-%d')]
                   for c in report['herd']['new_cows']])
    section_table(ws, 'Transfers', ['Cow', 'From block', 'To block', 'Date', 'Note'],
                  [[t.cow.tag_id, t.from_block.name if t.from_block else '', t.to_block.name,
                    t.transferred_at.strftime('%Y-%m-%d'), t.note] for t in report['herd']['transfers']])
    autosize(ws)

    ws = wb.create_sheet('Crops')
    section_table(ws, 'Activity log', ['Date', 'Crop', 'Activity', 'Harvested (kg)', 'Notes'],
                  [[a.date.strftime('%Y-%m-%d'), a.crop.name, a.get_activity_type_display(),
                    float(a.quantity_harvested_kg) if a.quantity_harvested_kg is not None else '', a.notes]
                   for a in report['crops']['activities']])
    autosize(ws)

    ws = wb.create_sheet('Inventory')
    section_table(ws, 'Stock movements', ['Date', 'Item', 'Type', 'Quantity', 'Note'],
                  [[m.date.strftime('%Y-%m-%d'), m.item.name, m.get_movement_type_display(),
                    float(m.quantity), m.note] for m in report['inventory']['movements']])
    section_table(ws, 'Current stock levels', ['Item', 'Category', 'Stock', 'Unit', 'Reorder level', 'Low stock?'],
                  [[i.name, i.get_category_display(), float(i.current_stock), i.get_unit_display(),
                    float(i.reorder_level), 'Yes' if i.is_low_stock else 'No']
                   for i in report['inventory']['levels']])
    autosize(ws)

    ws = wb.create_sheet('Finance')
    section_table(ws, 'By category', ['Kind', 'Category', 'Amount'],
                  [[r['kind'], r['category'], float(r['amount'])] for r in report['finance']['by_category']])
    section_table(ws, 'Transactions', ['Date', 'Kind', 'Category', 'Amount', 'Note'],
                  [[t.date.strftime('%Y-%m-%d'), t.get_kind_display(), t.get_category_display(),
                    float(t.amount), t.note] for t in report['finance']['transactions']])
    autosize(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_filename(report, "xlsx")}"'
    return response


# --------------------------------------------------------------------- PDF

def build_pdf_bytes(report):
    """The raw PDF bytes, reused both for the direct download (export_pdf)
    and as an email attachment (analysis:email_report)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm, leftMargin=16 * mm, rightMargin=16 * mm
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('FarmTitle', parent=styles['Title'], textColor=colors.HexColor('#065F46'))
    section_style = ParagraphStyle(
        'Section', parent=styles['Heading2'], textColor=colors.HexColor('#047857'), spaceBefore=14, spaceAfter=4
    )
    normal = styles['Normal']

    elements = []
    farm = report['farm']
    elements.append(Paragraph(f'{farm.name} ({farm.code})', title_style))
    elements.append(Paragraph(f'{report["period_label"]} Report', styles['Heading3']))
    elements.append(Paragraph(
        f'Period: {report["start_date"]} to {report["end_date"]} &nbsp;&nbsp;|&nbsp;&nbsp; '
        f'Generated {report["generated_at"].strftime("%d %b %Y %H:%M")}', normal
    ))
    elements.append(Spacer(1, 6 * mm))

    def add_table(headers, rows):
        data = [headers] + [[str(v) if v is not None else '-' for v in row] for row in rows]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#047857')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E7E5E4')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F4')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 5 * mm))

    elements.append(Paragraph('Milk Production', section_style))
    elements.append(Paragraph(
        f'Total: {report["milk"]["total_liters"]} L &nbsp;&nbsp; '
        f'Avg/active cow: {round(float(report["milk"]["avg_per_cow"]), 2)} L', normal
    ))
    add_table(['Block', 'Liters'], [[r['block__name'], r['liters']] for r in report['milk']['by_block']])
    add_table(['Tag', 'Name', 'Block', 'Liters'],
              [[r['cow__tag_id'], r['cow__name'], r['cow__block__name'], r['liters']] for r in report['milk']['by_cow']])
    add_table(['Date', 'Liters'], [[r['date'], r['liters']] for r in report['milk']['daily']])

    elements.append(Paragraph('Feeding', section_style))
    elements.append(Paragraph(
        f'Dairy meal: {report["feeding"]["total_dairy_meal_kg"]} kg &nbsp;&nbsp; '
        f'Silage/Hay: {report["feeding"]["total_silage_hay_kg"]} kg', normal
    ))
    add_table(['Block', 'Dairy meal (kg)', 'Silage/Hay (kg)'],
              [[r['block__name'], r['meal'] or 0, r['silage'] or 0] for r in report['feeding']['by_block']])

    elements.append(Paragraph('Herd', section_style))
    elements.append(Paragraph(
        f'Active cows: {report["herd"]["active_cows"]} &nbsp;&nbsp; '
        f'New cows: {len(report["herd"]["new_cows"])} &nbsp;&nbsp; '
        f'Transfers: {len(report["herd"]["transfers"])}', normal
    ))
    add_table(['Tag', 'Name', 'Block', 'Breed'],
              [[c.tag_id, c.name, c.block.name, c.breed] for c in report['herd']['new_cows']])
    add_table(['Cow', 'From', 'To', 'Date'],
              [[t.cow.tag_id, t.from_block.name if t.from_block else '-', t.to_block.name,
                t.transferred_at.strftime('%d %b %Y')] for t in report['herd']['transfers']])

    elements.append(Paragraph('Crops', section_style))
    elements.append(Paragraph(f'Total harvested: {report["crops"]["total_harvested_kg"]} kg', normal))
    add_table(['Date', 'Crop', 'Activity', 'Harvested (kg)'],
              [[a.date, a.crop.name, a.get_activity_type_display(), a.quantity_harvested_kg]
               for a in report['crops']['activities']])

    elements.append(Paragraph('Inventory', section_style))
    add_table(['Date', 'Item', 'Type', 'Quantity'],
              [[m.date, m.item.name, m.get_movement_type_display(), m.quantity]
               for m in report['inventory']['movements']])
    add_table(['Item', 'Stock', 'Unit', 'Low stock?'],
              [[i.name, i.current_stock, i.get_unit_display(), 'Yes' if i.is_low_stock else 'No']
               for i in report['inventory']['levels']])

    elements.append(Paragraph('Finance', section_style))
    elements.append(Paragraph(
        f'Income: {report["finance"]["total_income"]} &nbsp;&nbsp; '
        f'Expense: {report["finance"]["total_expense"]} &nbsp;&nbsp; '
        f'Net: {report["finance"]["net"]}', normal
    ))
    add_table(['Kind', 'Category', 'Amount'],
              [[r['kind'], r['category'], r['amount']] for r in report['finance']['by_category']])
    add_table(['Date', 'Kind', 'Category', 'Amount'],
              [[t.date, t.get_kind_display(), t.get_category_display(), t.amount]
               for t in report['finance']['transactions']])

    doc.build(elements)
    return buf.getvalue()


def export_pdf(report):
    response = HttpResponse(build_pdf_bytes(report), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_filename(report, "pdf")}"'
    return response
